import logging
from pathlib import Path
from typing import Dict, Any, Optional, List, Union
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from parsers.base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel files (xlsx, xls)."""
    
    def __init__(self, encoding: str = 'utf-8'):
        """
        Initialize ExcelParser.
        
        Args:
            encoding: Default encoding (not used for Excel but kept for compatibility)
        """
        super().__init__(encoding)
        self.sheets: Dict[str, pd.DataFrame] = {}
        self.active_sheet: Optional[str] = None
    
    def parse(self, file_path: Path, sheet_name: Optional[Union[str, int]] = None) -> pd.DataFrame:
        """
        Parse Excel file and return DataFrame.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to parse (None for active/first sheet)
            
        Returns:
            DataFrame containing Excel data
        """
        if not self.validate_file(file_path):
            return pd.DataFrame()
        
        try:
            # Determine file type and use appropriate engine
            if file_path.suffix.lower() == '.xls':
                engine = 'xlrd'
            else:
                engine = 'openpyxl'
            
            # Read all sheets first to understand structure
            excel_file = pd.ExcelFile(file_path, engine=engine)
            self.sheets = {}
            
            # If specific sheet requested
            if sheet_name is not None:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                self.sheets[str(sheet_name)] = df
                self.active_sheet = str(sheet_name)
                return self.clean_dataframe(df)
            
            # Parse all sheets
            for sheet in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet, engine=engine)
                    if not df.empty:
                        self.sheets[sheet] = self.clean_dataframe(df)
                        logger.debug(f"Parsed sheet '{sheet}': {len(df)} rows")
                except Exception as e:
                    logger.warning(f"Failed to parse sheet '{sheet}': {e}")
            
            # Return the largest sheet or first non-empty sheet
            if self.sheets:
                # Find sheet with most data
                largest_sheet = max(self.sheets.items(), 
                                  key=lambda x: len(x[1]) * len(x[1].columns))
                self.active_sheet = largest_sheet[0]
                logger.info(f"Selected sheet '{self.active_sheet}' as primary")
                return largest_sheet[1]
            
            return pd.DataFrame()
            
        except Exception as e:
            logger.error(f"Failed to parse Excel file {file_path}: {e}")
            return pd.DataFrame()
    
    def parse_all_sheets(self, file_path: Path) -> Dict[str, pd.DataFrame]:
        """
        Parse all sheets from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary mapping sheet names to DataFrames
        """
        if not self.validate_file(file_path):
            return {}
        
        try:
            # Parse if not already done
            if not self.sheets:
                self.parse(file_path)
            
            return self.sheets
            
        except Exception as e:
            logger.error(f"Failed to parse all sheets: {e}")
            return {}
    
    def extract_text(self, file_path: Path, max_chars: int = 50000) -> str:
        """
        Extract text content from Excel for LLM processing.
        
        Args:
            file_path: Path to the Excel file
            max_chars: Maximum number of characters to extract
            
        Returns:
            Extracted text content
        """
        if not self.validate_file(file_path):
            return ""
        
        text_parts = []
        total_chars = 0
        
        try:
            # Parse sheets if not already done
            if not self.sheets:
                self.parse_all_sheets(file_path)
            
            for sheet_name, df in self.sheets.items():
                if total_chars >= max_chars:
                    break
                
                # Add sheet header
                sheet_text = f"\n=== Sheet: {sheet_name} ===\n"
                
                # Add column names
                sheet_text += "Columns: " + ", ".join(df.columns.astype(str)) + "\n\n"
                
                # Add data rows (limit to prevent huge text)
                max_rows = min(100, len(df))
                for idx, row in df.head(max_rows).iterrows():
                    if total_chars + len(sheet_text) >= max_chars:
                        break
                    
                    row_text = " | ".join(str(val) for val in row.values)
                    sheet_text += f"Row {idx + 1}: {row_text}\n"
                
                # Add summary if more rows exist
                if len(df) > max_rows:
                    sheet_text += f"\n... and {len(df) - max_rows} more rows\n"
                
                text_parts.append(sheet_text)
                total_chars += len(sheet_text)
            
            return "".join(text_parts)[:max_chars]
            
        except Exception as e:
            logger.error(f"Failed to extract text from Excel: {e}")
            return ""
    
    def get_metadata(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract metadata from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing Excel metadata
        """
        metadata = super().get_metadata(file_path)
        
        try:
            # Basic metadata from pandas
            excel_file = pd.ExcelFile(file_path)
            metadata.update({
                "sheet_count": len(excel_file.sheet_names),
                "sheet_names": excel_file.sheet_names,
                "active_sheet": self.active_sheet
            })
            
            # Try to get more metadata using openpyxl for xlsx files
            if file_path.suffix.lower() == '.xlsx':
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # Document properties
                props = workbook.properties
                if props:
                    metadata["excel_metadata"] = {
                        "title": props.title or "",
                        "author": props.creator or "",
                        "created": str(props.created) if props.created else "",
                        "modified": str(props.modified) if props.modified else "",
                        "last_modified_by": props.lastModifiedBy or "",
                        "category": props.category or "",
                        "comments": props.description or ""
                    }
                
                # Sheet details
                sheet_details = {}
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_details[sheet_name] = {
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "dimension": str(sheet.dimensions) if sheet.dimensions else None
                    }
                
                metadata["sheet_details"] = sheet_details
                workbook.close()
            
            # Add data statistics if sheets are parsed
            if self.sheets:
                metadata["data_stats"] = {}
                for sheet_name, df in self.sheets.items():
                    metadata["data_stats"][sheet_name] = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "null_values": int(df.isnull().sum().sum()),
                        "numeric_columns": len(df.select_dtypes(include=['number']).columns),
                        "text_columns": len(df.select_dtypes(include=['object']).columns)
                    }
            
        except Exception as e:
            logger.debug(f"Could not extract Excel metadata: {e}")
        
        return metadata
    
    def detect_header_row(self, file_path: Path, sheet_name: Optional[str] = None) -> int:
        """
        Detect the header row in an Excel sheet.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to analyze
            
        Returns:
            Row index of likely header (0-based)
        """
        try:
            # Read first 20 rows without header
            df_sample = pd.read_excel(
                file_path, 
                sheet_name=sheet_name,
                header=None,
                nrows=20
            )
            
            # Look for row with most non-null, unique values
            best_row = 0
            best_score = 0
            
            for i in range(min(10, len(df_sample))):
                row = df_sample.iloc[i]
                
                # Skip if row is mostly empty
                if row.isna().sum() > len(row) * 0.7:
                    continue
                
                # Calculate score based on uniqueness and non-null values
                non_null = row.notna().sum()
                unique = row.nunique()
                
                # Check if values look like headers (strings, not numbers)
                string_values = sum(1 for val in row if isinstance(val, str))
                
                score = non_null + unique * 2 + string_values * 3
                
                if score > best_score:
                    best_score = score
                    best_row = i
            
            return best_row
            
        except Exception as e:
            logger.debug(f"Could not detect header row: {e}")
            return 0
    
    def extract_formulas(self, file_path: Path, sheet_name: Optional[str] = None) -> Dict[str, str]:
        """
        Extract formulas from Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to extract from
            
        Returns:
            Dictionary mapping cell references to formulas
        """
        formulas = {}
        
        if file_path.suffix.lower() != '.xlsx':
            return formulas
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            sheets_to_check = [sheet_name] if sheet_name else workbook.sheetnames
            
            for sheet_name in sheets_to_check:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            cell_ref = f"{sheet_name}!{cell.coordinate}"
                            formulas[cell_ref] = cell.value
            
            workbook.close()
            
        except Exception as e:
            logger.debug(f"Could not extract formulas: {e}")
        
        return formulas
    
    def get_cell_formatting(self, file_path: Path, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Extract cell formatting information.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to analyze
            
        Returns:
            Dictionary with formatting information
        """
        formatting = {
            "merged_cells": [],
            "colored_cells": [],
            "bold_cells": [],
            "borders": []
        }
        
        if file_path.suffix.lower() != '.xlsx':
            return formatting
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            
            # Merged cells
            for merged_range in sheet.merged_cells.ranges:
                formatting["merged_cells"].append(str(merged_range))
            
            # Check formatting in first 100 cells
            for row in sheet.iter_rows(max_row=20, max_col=20):
                for cell in row:
                    if cell.value is not None:
                        # Check for fill color
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            formatting["colored_cells"].append(cell.coordinate)
                        
                        # Check for bold
                        if cell.font and cell.font.bold:
                            formatting["bold_cells"].append(cell.coordinate)
                        
                        # Check for borders
                        if cell.border and any([
                            cell.border.top, cell.border.bottom,
                            cell.border.left, cell.border.right
                        ]):
                            formatting["borders"].append(cell.coordinate)
            
            workbook.close()
            
        except Exception as e:
            logger.debug(f"Could not extract formatting: {e}")
        
        return formatting